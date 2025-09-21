# processor.py — GPT-разбор инвойса и генерация GOST (ST00012) QR
# v2.9:
# - Понятные русские причины отказа (а не "PersonalAcc must be 20 digits").
# - Предпросмотр распознанных полей в fallback-сообщении.
# - Автоповтор с более сильной моделью (RETRY_MODEL=gpt-4o), если валидация провалена.
# - OCR: рендер PDF-сканов при 360 DPI (PyMuPDF), фиксы цифр O→0, I/l→1 и т.д.
# - Excel: авто-детект xlsx/xls/csv; DOCX поддержан; строгая санитарка и пересборка ST00012.

from __future__ import annotations
import io
import os
import re
import json
import csv
import base64
import logging
from typing import Tuple, Optional, List

from openai import OpenAI
import qrcode  # pillow обязателен

from telegram.ext import ContextTypes
from store import store

log = logging.getLogger("processor")

NBSP = "\u00A0"
ST00012_REQUIRED = ["Name", "PersonalAcc", "BankName", "BIC", "CorrespAcc", "Sum", "Purpose"]
OPTIONAL_FIELDS = ["PayeeINN", "KPP"]

# ---------------- настройки моделей ----------------
GPT_MODEL = os.getenv("GPT_INVOICE_MODEL", "gpt-4o-mini")
RETRY_MODEL = os.getenv("GPT_RETRY_MODEL", "gpt-4o")  # усиленная модель для автоповтора
MAX_RETRY_ON_FAIL = int(os.getenv("GPT_MAX_RETRY_ON_FAIL", "1"))

# ---------------- utils ----------------
def _qr_png_bytes(payload: str) -> bytes:
    img = qrcode.make(payload)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def _to_data_uri(b: bytes, mime: str) -> str:
    return f"data:{mime};base64,{base64.b64encode(b).decode('ascii')}"

def _guess_mime_for_photo() -> str:
    return "image/jpeg"

def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def _ocr_digit_fix(s: str) -> str:
    """Подправляем частые OCR-ошибки в числовых строках: O→0, o→0, I/l→1, B→8, S→5, Z→2."""
    if not isinstance(s, str):
        return s
    table = str.maketrans({
        "O": "0", "o": "0",
        "I": "1", "l": "1", "í": "1",
        "B": "8",
        "S": "5",
        "Z": "2"
    })
    return s.translate(table)

def _sanitize_fields(fields: dict) -> dict:
    f = dict(fields or {})
    for k in ["PersonalAcc", "CorrespAcc", "BIC", "Sum", "PayeeINN", "KPP"]:
        if k in f and isinstance(f[k], str):
            f[k] = _ocr_digit_fix(f[k])
            f[k] = _digits_only(f[k])
    if "Purpose" in f and isinstance(f["Purpose"], str):
        p = f["Purpose"].replace(NBSP, " ").strip()
        p = re.sub(r"\s+", " ", p)
        p = p.replace("«", '"').replace("»", '"')
        f["Purpose"] = p
    for k in ["Name", "BankName"]:
        if k in f and isinstance(f[k], str):
            s = f[k].replace(NBSP, " ").replace("«", '"').replace("»", '"').strip()
            s = re.sub(r"\s+", " ", s)
            f[k] = s
    return f

def _build_st00012_from_fields(fields: dict) -> str:
    parts = [
        f"Name={fields.get('Name','')}",
        f"PersonalAcc={fields.get('PersonalAcc','')}",
        f"BankName={fields.get('BankName','')}",
        f"BIC={fields.get('BIC','')}",
        f"CorrespAcc={fields.get('CorrespAcc','')}",
        f"Sum={fields.get('Sum','')}",
        f"Purpose={fields.get('Purpose','')}",
    ]
    if fields.get("PayeeINN"):
        parts.append(f"PayeeINN={fields['PayeeINN']}")
    if fields.get("KPP"):
        parts.append(f"KPP={fields['KPP']}")
    return "ST00012|" + "|".join(parts)

def _fields_preview(fields: dict) -> str:
    show = []
    def take(k, title):
        v = fields.get(k)
        if v is not None:
            v = str(v)
            if len(v) > 120: v = v[:117] + "..."
            show.append(f"{title}: {v}")
    take("Name", "Получатель")
    take("PayeeINN", "ИНН")
    take("KPP", "КПП")
    take("BankName", "Банк")
    take("BIC", "БИК")
    take("CorrespAcc", "Корр. счёт")
    take("PersonalAcc", "Р/счёт")
    if "Sum" in fields:
        try:
            rub = f"{int(fields['Sum'])/100:.2f}"
        except Exception:
            rub = str(fields['Sum'])
        show.append(f"Сумма: {rub} RUB (в копейках: {fields.get('Sum')})")
    take("Purpose", "Назначение")
    return "\n".join(show)

# ---------------- signatures/helpers ----------------
def _is_pdf(b: bytes) -> bool:
    return b[:4] == b"%PDF"

def _is_xlsx_zip(b: bytes) -> bool:
    return b[:4] == b"PK\x03\x04"

def _is_xls_ole(b: bytes) -> bool:
    return b[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

# ---------------- local extractors ----------------
INV_NUM_RE  = re.compile(r"(?:Сч[её]т(?:\s*на\s*оплату)?\s*№\s*([0-9\-]+))", re.IGNORECASE)
INV_DATE_RE = re.compile(r"от\s*([0-9]{1,2}[.\s][0-9]{1,2}[.\s][0-9]{2,4}|[0-9]{1,2}\s+[А-Яа-яЁёA-Za-z]+?\s+\d{4})")
VAT_PCT_RE  = re.compile(r"(?:НДС|VAT)\s*([0-9]{1,2})\s*%")
VAT_SUM_RE  = re.compile(r"(?:НДС[:\s]|VAT[:\s]).{0,30}?([0-9\s\u00A0]+(?:[.,][0-9]{1,2})?)", re.IGNORECASE)
TOTAL_RE    = re.compile(r"(?:Всего\s*к\s*оплате|Итого|Total).{0,30}?([0-9\s\u00A0]+(?:[.,][0-9]{1,2})?)", re.IGNORECASE)

def _pdf_to_text(file_bytes: bytes) -> str:
    try:
        from PyPDF2 import PdfReader
        r = PdfReader(io.BytesIO(file_bytes))
        chunks = []
        for p in r.pages:
            try:
                chunks.append((p.extract_text() or "").replace(NBSP, " "))
            except Exception:
                pass
        return "\n".join(chunks)
    except Exception as e:
        log.warning("PDF extract failed: %s", e)
        return ""

def _pdf_to_images(file_bytes: bytes, max_pages: int = 3, dpi: int = 360) -> List[bytes]:
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        log.warning("PyMuPDF not available: %s", e)
        return []
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception as e:
        log.warning("fitz.open failed: %s", e)
        return []
    images = []
    try:
        pages = min(len(doc), max_pages)
        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)
        for i in range(pages):
            page = doc.load_page(i)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            images.append(pix.tobytes("png"))
    except Exception as e:
        log.warning("PDF render to images failed: %s", e)
    finally:
        doc.close()
    return images

def _xls_to_text(file_bytes: bytes) -> str:
    try:
        import xlrd  # type: ignore
        book = xlrd.open_workbook(file_contents=file_bytes)
        out = []
        for si in range(book.nsheets):
            sh = book.sheet_by_index(si)
            for ri in range(sh.nrows):
                row = sh.row_values(ri)
                vals = [str(v) if v is not None else "" for v in row]
                if any(vals):
                    out.append(" | ".join(vals))
        return "\n".join(out)
    except Exception as e:
        log.warning("XLS extract failed: %s", e)
        return ""

def _xlsx_to_text(file_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) if v is not None else "" for v in row]
                if any(vals):
                    out.append(" | ".join(vals))
        return "\n".join(out)
    except Exception as e:
        log.warning("XLSX extract failed: %s", e)
        return ""

def _csv_like_to_text(file_bytes: bytes) -> str:
    encodings = ("utf-8", "cp1251", "latin-1")
    text = ""
    for enc in encodings:
        try:
            text = file_bytes.decode(enc, errors="strict")
            break
        except Exception:
            continue
    if not text:
        text = file_bytes.decode("utf-8", errors="ignore")
    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(text[:4096])
        reader = csv.reader(text.splitlines(), dialect)
        rows = [" | ".join(row) for row in reader if any(cell.strip() for cell in row)]
        return "\n".join(rows)
    except Exception:
        return text

def _excel_to_text(file_bytes: bytes) -> str:
    if _is_xlsx_zip(file_bytes):
        txt = _xlsx_to_text(file_bytes)
        if txt:
            return txt
    if _is_xls_ole(file_bytes):
        txt = _xls_to_text(file_bytes)
        if txt:
            return txt
    return _csv_like_to_text(file_bytes)

def _docx_to_text(file_bytes: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        log.warning("DOCX extract failed: %s", e)
        return ""

def _normalize_money(s: str) -> Optional[float]:
    if not s:
        return None
    s = s.replace(NBSP, " ").replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        try:
            return float(s.replace(".", ""))
        except Exception:
            return None

def _pre_hint(text: str) -> dict:
    t = (text or "").replace(NBSP, " ")
    inv_num = inv_date = None
    vat_pct = vat_sum = total = None

    m = INV_NUM_RE.search(t)
    if m: inv_num = m.group(1).strip()

    m = INV_DATE_RE.search(t)
    if m: inv_date = m.group(1).strip()

    m = VAT_PCT_RE.search(t)
    if m:
        try: vat_pct = int(m.group(1))
        except Exception: pass

    m = VAT_SUM_RE.search(t)
    if m:
        v = _normalize_money(m.group(1))
        if v is not None: vat_sum = v

    m = TOTAL_RE.search(t)
    if m:
        v = _normalize_money(m.group(1))
        if v is not None: total = v

    return {
        "invoice_number": inv_num,
        "invoice_date": inv_date,
        "vat_percent": vat_pct,
        "vat_amount": vat_sum,
        "total": total,
    }

# ---------------- validate & caption ----------------
def _validate_st00012(st: str) -> Optional[str]:
    if not st or not st.startswith("ST00012|"):
        return "payload is not ST00012"
    fields = {}
    try:
        for p in st.split("|")[1:]:
            if "=" in p:
                k, v = p.split("=", 1)
                fields[k] = v
    except Exception:
        return "failed to parse key=value pairs"

    missing = [k for k in ST00012_REQUIRED if k not in fields or not str(fields[k]).strip()]
    if missing:
        return "missing:" + ",".join(missing)

    bic = re.sub(r"\D+", "", fields["BIC"])
    pa  = re.sub(r"\D+", "", fields["PersonalAcc"])
    ca  = re.sub(r"\D+", "", fields["CorrespAcc"])
    s   = re.sub(r"\D+", "", fields["Sum"])

    if len(bic) != 9:
        return "bad_bic"
    if len(pa) != 20:
        return "bad_personal"
    if len(ca) != 20:
        return "bad_corresp"
    if not s.isdigit() or int(s) <= 0:
        return "bad_sum"

    purpose = (fields.get("Purpose") or "").strip()
    if not purpose:
        return "bad_purpose"

    return None

def _reason_human(code: str, st_fields: dict | None, fields_src: dict | None) -> str:
    # дружелюбное описание причины на русском
    if not code:
        return ""
    if code.startswith("missing:"):
        miss = code.split(":", 1)[1].split(",")
        names = {
            "Name": "Получатель (Name)",
            "PersonalAcc": "Р/счёт (PersonalAcc)",
            "BankName": "Банк (BankName)",
            "BIC": "БИК (BIC)",
            "CorrespAcc": "Корр. счёт (CorrespAcc)",
            "Sum": "Сумма (в копейках)",
            "Purpose": "Назначение платежа",
        }
        items = ", ".join(names.get(m, m) for m in miss)
        return f"Отсутствуют обязательные поля: {items}."
    if code == "bad_bic":
        return "БИК должен содержать ровно 9 цифр."
    if code == "bad_personal":
        return "Р/счёт получателя должен содержать ровно 20 цифр."
    if code == "bad_corresp":
        return "Корреспондентский счёт должен содержать ровно 20 цифр."
    if code == "bad_sum":
        return "Сумма должна быть положительным целым числом в копейках."
    if code == "bad_purpose":
        return "Назначение платежа не должно быть пустым."
    if code == "payload is not ST00012":
        return "Строка QR не соответствует формату ST00012."
    if code == "failed to parse key=value pairs":
        return "Не удалось распарсить пары ключ=значение в ST00012."
    return code

def _caption_from_fields(fields: dict, notes: str = "") -> str:
    name = fields.get("Name", "")
    sum_kopecks = fields.get("Sum", "0")
    try:
        amount = f"{int(sum_kopecks)/100:.2f}"
    except Exception:
        amount = "0.00"
    purpose = fields.get("Purpose", "")
    extra = []
    if fields.get("PayeeINN"):
        extra.append(f"ИНН: {fields['PayeeINN']}")
    if fields.get("KPP"):
        extra.append(f"КПП: {fields['KPP']}")
    if notes:
        # укорачиваем заметки, чтобы не раздувать подпись
        n = notes.strip()
        if len(n) > 300: n = n[:297] + "..."
        extra.append(f"Примечание: {n}")
    tail = ("\n" + "\n".join(extra)) if extra else ""
    return f"Получатель: {name}\nСумма: {amount} RUB\nНазначение: {purpose}{tail}"

# ---------------- GPT ----------------
SYSTEM_PROMPT = (
    "Ты финансовый парсер инвойсов. По входному контенту (текст PDF/таблицы/документа или изображения страниц) "
    "найди реквизиты для платежного QR по стандарту GOST ST00012 (Россия). Отвечай строго JSON-объектом: "
    '{"st":"ST00012|Name=...|PersonalAcc=...|BankName=...|BIC=...|CorrespAcc=...|Sum=...|Purpose=...",'
    '"fields":{"Name":"...","PersonalAcc":"...","BankName":"...","BIC":"...","CorrespAcc":"...","Sum":"...",'
    '"Purpose":"...","PayeeINN":"(если есть)","KPP":"(если есть)"},'
    '"notes":"..."} '
    "Требования: Sum — целое число копеек (например, 179500 для 1 795,00). "
    "Purpose обязательно: если есть НДС — укажи «НДС X% — Y ₽», если нет — «без НДС». "
    "Если видишь номер и дату счёта, добавь «Оплата по счёту №… от …». "
    "Не выдумывай данные: если чего-то нет в документе — оставь пустым и распиши в notes. "
    "Все номера счетов/БИК должны быть выведены цифрами без пробелов, длины: PersonalAcc=20, CorrespAcc=20, BIC=9."
)

def _client() -> OpenAI:
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is missing")
    return OpenAI(api_key=api_key)

def _parse_json(text: str) -> dict:
    s, e = text.find("{"), text.rfind("}")
    if s == -1 or e == -1 or e <= s:
        raise ValueError("no JSON object found")
    return json.loads(text[s:e+1])

async def _call_gpt_on_file(
    file_bytes: bytes,
    file_type: str,
    prehint: dict,
    docx_text: str = "",
    model: Optional[str] = None
) -> Tuple[Optional[str], Optional[dict], Optional[str]]:
    client = _client()
    mdl = model or GPT_MODEL

    if file_type == "photo":
        mime = _guess_mime_for_photo()
        data_uri = _to_data_uri(file_bytes, mime)
        user_content = [
            {"type": "text", "text": "Извлеки реквизиты по изображению счёта и верни JSON как описано."},
            {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
            {"type": "image_url", "image_url": {"url": data_uri}},
        ]
    elif file_type == "document":
        txt = _pdf_to_text(file_bytes) if _is_pdf(file_bytes) else ""
        if txt and txt.strip():
            user_content = [
                {"type": "text", "text": "Ниже текст документа (PDF). Верни JSON как описано."},
                {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
                {"type": "text", "text": txt[:15000]},
            ]
        else:
            if _is_pdf(file_bytes):
                images = _pdf_to_images(file_bytes, max_pages=3, dpi=360)
                if not images:
                    return None, None, "PDF is a scan and could not be rendered to images"
                user_content = [
                    {"type": "text", "text": "PDF выглядит как скан. Проанализируй изображения страниц и верни JSON как описано."},
                    {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
                ]
                for img in images:
                    user_content.append({"type": "image_url", "image_url": {"url": _to_data_uri(img, "image/png")}})
            else:
                # если это не PDF-документ, попробуем как DOCX-текст
                if docx_text and docx_text.strip():
                    user_content = [
                        {"type": "text", "text": "Ниже текст из DOCX. Верни JSON как описано."},
                        {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
                        {"type": "text", "text": docx_text[:15000]},
                    ]
                else:
                    user_content = [
                        {"type": "text", "text": "Текст документа не извлечён. Верни JSON как описано, если возможно."},
                        {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
                    ]
    elif file_type == "excel":
        txt = _excel_to_text(file_bytes)
        user_content = [
            {"type": "text", "text": "Ниже текстовое представление Excel/CSV-счёта. Верни JSON как описано."},
            {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
            {"type": "text", "text": txt[:15000] if txt else ""},
        ]
    else:
        txt = _pdf_to_text(file_bytes)
        user_content = [
            {"type": "text", "text": "Ниже текст из документа. Верни JSON как описано."},
            {"type": "text", "text": f"Подсказки (если релевантны): {json.dumps(prehint, ensure_ascii=False)}"},
            {"type": "text", "text": txt[:15000] if txt else ""},
        ]

    try:
        resp = client.chat.completions.create(
            model=mdl,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            temperature=0.0,
        )
        raw = resp.choices[0].message.content or ""
    except Exception as e:
        return None, None, f"GPT error: {e}"

    try:
        data = _parse_json(raw)
    except Exception as e:
        return None, None, f"Bad JSON from GPT: {e}"

    st = (data.get("st") or "").strip()
    fields = data.get("fields") or {}
    notes = data.get("notes") or ""
    return st, fields, notes

# ---------------- main entry ----------------
async def on_approved_send_qr(context: ContextTypes.DEFAULT_TYPE, *, chat_id: int, status_msg_id: int) -> None:
    inv = store.get(status_msg_id)
    if not inv or not inv.get("src"):
        log.warning("No source bound to status_msg_id=%s", status_msg_id)
        return

    src = inv["src"]
    file_id = src["file_id"]
    file_type = src["file_type"]   # "document" | "photo" | "excel"
    thread_id = src.get("thread_id")

    tg_file = await context.bot.get_file(file_id)
    fb = await tg_file.download_as_bytearray()
    b = bytes(fb)

    base_text = ""
    docx_text = ""
    if file_type == "document":
        if _is_pdf(b):
            base_text = _pdf_to_text(b)
        else:
            docx_text = _docx_to_text(b)
            base_text = docx_text
    elif file_type == "excel":
        base_text = _excel_to_text(b)
    prehint = _pre_hint(base_text)

    # 1-я попытка: текущая модель
    st, fields, notes = await _call_gpt_on_file(b, file_type, prehint, docx_text=docx_text, model=GPT_MODEL)

    # Санитарка + фиксы цифр
    if fields:
        fields = _sanitize_fields(fields)

    # Пересборка st, если пустой/кривой
    if (not st or not st.startswith("ST00012|")) and fields:
        if "Sum" in fields and fields["Sum"] and not fields["Sum"].isdigit():
            rub = re.sub(r"[^\d,\.]", "", fields["Sum"]).replace(",", ".")
            try:
                fields["Sum"] = str(int(round(float(rub) * 100)))
            except Exception:
                pass
        st = _build_st00012_from_fields(fields)

    # Проверка
    err_code = _validate_st00012(st) if st else "payload is not ST00012"

    # Автоповтор с более сильной моделью, если надо
    retries_left = MAX_RETRY_ON_FAIL if GPT_MODEL != RETRY_MODEL else 0
    while err_code and retries_left > 0:
        retries_left -= 1
        st2, fields2, notes2 = await _call_gpt_on_file(b, file_type, prehint, docx_text=docx_text, model=RETRY_MODEL)
        if fields2:
            fields2 = _sanitize_fields(fields2)
        if (not st2 or not st2.startswith("ST00012|")) and fields2:
            if "Sum" in fields2 and fields2["Sum"] and not fields2["Sum"].isdigit():
                rub = re.sub(r"[^\d,\.]", "", fields2["Sum"]).replace(",", ".")
                try:
                    fields2["Sum"] = str(int(round(float(rub) * 100)))
                except Exception:
                    pass
            st2 = _build_st00012_from_fields(fields2)
        err2 = _validate_st00012(st2) if st2 else "payload is not ST00012"
        # Если новая попытка лучше — используем её
        if not err2 or (err2 and st2 and fields2 and len(_fields_preview(fields2)) > len(_fields_preview(fields or {}))):
            st, fields, notes, err_code = st2, fields2, notes2, err2
            break

    # Успех → отправляем QR
    if not err_code and st and fields:
        try:
            png = _qr_png_bytes(st)
            caption = _caption_from_fields(fields, notes=notes)
            await context.bot.send_photo(
                chat_id=chat_id,
                photo=png,
                caption=caption,
                message_thread_id=thread_id if thread_id else None,
                reply_to_message_id=status_msg_id,
            )
            return
        except Exception as e:
            err_code = f"Не удалось сгенерировать QR: {e}"

    # Fallback c русским объяснением и предпросмотром полей
    reason = _reason_human(err_code, None, fields)
    preview = _fields_preview(fields or {})
    reason_block = f"\nПричина: {reason}" if reason else ""
    preview_block = f"\n\nРаспознанные поля (проверьте внимательнее):\n{preview}" if preview else ""

    demo_payload = "ST00012|Name=ERROR|PersonalAcc=00000000000000000000|BankName=ERROR|BIC=000000000|CorrespAcc=00000000000000000000|Sum=0|Purpose=Parse failed"
    png = _qr_png_bytes(demo_payload)
    fallback_caption = (
        "Не удалось собрать рабочий QR. Проверьте реквизиты или пришлите более качественный образец для настройки."
        + reason_block + preview_block
    )

    await context.bot.send_photo(
        chat_id=chat_id,
        photo=png,
        caption=fallback_caption,
        message_thread_id=thread_id if thread_id else None,
        reply_to_message_id=status_msg_id,
    )
